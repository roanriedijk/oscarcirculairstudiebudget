"""
Studiebudget-calculator — Oscar Circulair
Streamlit-versie.

Berekent per salarisperiode het meetellende brutoloon binnen een
spending window van 2 jaar (gerekend vanaf de laatste contract-einddatum)
en daarvan 2% als studiebudget. Splitst op in reeds opgebouwd (t/m vandaag)
en vooraf op te nemen (toekomstig contractloon).
"""

import io
import calendar
from datetime import date, timedelta

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RATE = 0.02

# ---- Kleuren (huisstijl) -------------------------------------------------
BLUE = "1B3F8F"
BLUE_DEEP = "13306E"
GOLD = "908830"
LIGHT = "F3F6FC"
GREY = "8A93A8"
INK = "1A2233"


# ---- Hulpfuncties --------------------------------------------------------
def fmt_eur(n: float) -> str:
    """Bedrag als € 1.234,56 (Nederlandse notatie)."""
    if n is None or n != n:  # None of NaN
        n = 0.0
    s = f"{n:,.2f}"  # 1,234.56
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"€ {s}"


def fmt_date(d: date) -> str:
    maanden = ["jan", "feb", "mrt", "apr", "mei", "jun",
               "jul", "aug", "sep", "okt", "nov", "dec"]
    return f"{d.day} {maanden[d.month - 1]} {d.year}"


def days_in_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def window_start_for(end_date: date) -> date:
    """Het spending window loopt 2 jaar terug vanaf de contract-einddatum."""
    try:
        return end_date.replace(year=end_date.year - 2)
    except ValueError:  # 29 feb
        return end_date.replace(year=end_date.year - 2, day=28)


def gross_for_range(start: date, end: date, monthly: float) -> float:
    """
    Som het brutoloon over [start, end], waarbij elke kalendermaand
    pro-rata wordt meegeteld op basis van het aantal dagen dat binnen
    het bereik valt.
    """
    if not start or not end or end < start:
        return 0.0
    total = 0.0
    y, m = start.year, start.month
    while date(y, m, 1) <= end:
        dim = days_in_month(y, m)
        month_start = date(y, m, 1)
        month_end = date(y, m, dim)
        overlap_start = max(month_start, start)
        overlap_end = min(month_end, end)
        if overlap_end >= overlap_start:
            days = (overlap_end - overlap_start).days + 1
            total += monthly * (days / dim)
        # volgende maand
        if m == 12:
            y, m = y + 1, 1
        else:
            m += 1
    return total


def clamp_to(start, end, win_start, win_end):
    """Knip een bereik bij tot [win_start, win_end]. None als geen overlap."""
    s = max(start, win_start)
    e = min(end, win_end)
    if e < s:
        return None
    return (s, e)


def up_to(start, end, cap):
    """Snijd een bereik met alles t/m cap. None als niets overblijft."""
    if not start or not end:
        return None
    e = min(end, cap)
    if e < start:
        return None
    return (start, e)


# ---- Pagina-instellingen -------------------------------------------------
st.set_page_config(page_title="Studiebudget-calculator", page_icon="🎓", layout="centered")

st.markdown(
    f"""
    <style>
    .stApp {{ background:#f4f6fb; }}
    /* Force readable dark text regardless of the visitor's dark/light setting */
    .stApp, .stApp p, .stApp label, .stApp span, .stApp li,
    .stApp h1, .stApp h2, .stApp h3, .stApp h4,
    [data-testid="stMarkdownContainer"], [data-testid="stWidgetLabel"],
    [data-testid="stCaptionContainer"] {{
        color:#{INK} !important;
    }}
    .ocs-eyebrow {{
        font-size:12px; letter-spacing:0.18em; text-transform:uppercase;
        color:#{GOLD} !important; font-weight:700; margin-bottom:4px;
    }}
    .ocs-title {{
        font-size:32px; font-weight:800; color:#{BLUE_DEEP} !important;
        line-height:1.1; margin:0 0 8px 0;
    }}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown('<div class="ocs-eyebrow">HR · Studiebudget</div>', unsafe_allow_html=True)
st.markdown('<div class="ocs-title">Studiebudget-calculator</div>', unsafe_allow_html=True)

st.write(
    "Voer per periode het **bruto maandsalaris** in. De app berekent het totale "
    "brutoloon over die periode en daarvan **2%** als studiebudget. Het budget mag "
    "besteed worden over een doorlopende periode van **twee jaar**, gerekend vanaf de "
    "**einddatum van het contract** (de laatste einddatum die je invult). Loon van vóór "
    "dat venster vervalt; toekomstig loon binnen het lopende contract telt wél mee."
)

today = date.today()

# ---- Invoer: naam --------------------------------------------------------
name = st.text_input("Naam medewerker (optioneel)", placeholder="bijv. Sanne de Vries")

# ---- Invoer: perioden via data-editor ------------------------------------
st.subheader("Salarisperioden")
st.caption("Voeg een rij toe met de **+** onderaan de tabel. Vul start- en einddatum "
           "in (JJJJ-MM-DD) en het bruto maandsalaris.")

if "periods_df" not in st.session_state:
    st.session_state.periods_df = pd.DataFrame(
        [{"Startdatum": None, "Einddatum": None, "Bruto maandsalaris (€)": None}]
    )

edited = st.data_editor(
    st.session_state.periods_df,
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        "Startdatum": st.column_config.DateColumn("Startdatum", format="YYYY-MM-DD"),
        "Einddatum": st.column_config.DateColumn("Einddatum", format="YYYY-MM-DD"),
        "Bruto maandsalaris (€)": st.column_config.NumberColumn(
            "Bruto maandsalaris (€)", min_value=0.0, step=50.0, format="%.2f"
        ),
    },
    key="editor",
)

already_used = st.number_input("Al gebruikt (€, optioneel)", min_value=0.0, value=0.0, step=50.0)

calculate = st.button("Uitrekenen", type="primary")


# ---- Berekening ----------------------------------------------------------
def to_date(v):
    if v is None or (isinstance(v, float) and v != v):
        return None
    if isinstance(v, date):
        return v
    try:
        return pd.to_datetime(v).date()
    except Exception:
        return None


def build_rows(df):
    """Verwerk de ingevoerde rijen naar berekende regels."""
    raw = []
    for _, row in df.iterrows():
        start = to_date(row.get("Startdatum"))
        end = to_date(row.get("Einddatum"))
        monthly = row.get("Bruto maandsalaris (€)")
        monthly = float(monthly) if monthly not in (None, "") and monthly == monthly else 0.0
        # negeer volledig lege rijen
        if start is None and end is None and not monthly:
            continue
        raw.append({"start": start, "end": end, "monthly": monthly})

    # contract-einddatum = laatste einddatum
    contract_end = None
    for r in raw:
        if r["end"] and (contract_end is None or r["end"] > contract_end):
            contract_end = r["end"]

    window_start = window_start_for(contract_end) if contract_end else None

    results = []
    for r in raw:
        start, end, monthly = r["start"], r["end"], r["monthly"]
        note = None
        eligible_gross = past_gross = future_gross = full_gross = 0.0
        clamped = None

        if start and end and end >= start and window_start:
            full_gross = gross_for_range(start, end, monthly)
            clamped = clamp_to(start, end, window_start, contract_end)
            if not clamped:
                note = "Buiten de periode van 2 jaar — telt niet mee"
            else:
                eligible_gross = gross_for_range(clamped[0], clamped[1], monthly)
                past_part = up_to(clamped[0], clamped[1], today)
                if past_part:
                    past_gross = gross_for_range(past_part[0], past_part[1], monthly)
                future_gross = eligible_gross - past_gross
                if clamped[0] > start:
                    note = f"Deels vóór de geldige periode — meegerekend vanaf {fmt_date(clamped[0])}"
        elif start and end and end < start:
            note = "Einddatum ligt vóór de startdatum"
        elif start and not end and monthly > 0:
            note = "Vul een einddatum in om deze regel mee te rekenen"

        results.append({
            "start": start, "end": end, "monthly": monthly,
            "full_gross": full_gross, "eligible_gross": eligible_gross,
            "past_gross": past_gross, "future_gross": future_gross,
            "budget": eligible_gross * RATE, "clamped": clamped, "note": note,
        })

    return results, contract_end, window_start


def make_excel(rows, contract_end, window_start, total_eligible, total_budget,
               accrued, advance, used, remaining, name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Studiebudget"
    ws.sheet_view.showGridLines = False

    thin_blue = Side(style="thin", color=BLUE_DEEP)
    hair = Side(style="hair", color="D3DBEB")
    euro_fmt = '€ #,##0.00'

    widths = [18, 18, 20, 34, 22, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Titel
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "Studiebudget"
    c.font = Font(name="Arial", size=20, bold=True, color=BLUE_DEEP)
    c.alignment = Alignment(vertical="center")

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "Oscar Circulair · HR"
    c.font = Font(name="Arial", size=11, bold=True, color=GOLD)

    # Meta
    r = 4
    meta = [
        ("Medewerker", name or "—"),
        ("Geldige periode",
         f"{fmt_date(window_start)} t/m {fmt_date(contract_end)}" if window_start and contract_end else "—"),
        ("Peildatum (vandaag)", fmt_date(today)),
    ]
    for k, v in meta:
        kc = ws.cell(row=r, column=1, value=k)
        kc.font = Font(name="Arial", size=10, bold=True, color=GREY)
        vc = ws.cell(row=r, column=2, value=v)
        vc.font = Font(name="Arial", size=11, color=INK)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    r += 1

    # Tabelkop
    headers = ["Startdatum", "Einddatum", "Bruto maandsalaris",
               "Meegerekende periode", "Meetellend brutoloon", "Studiebudget (2%)"]
    header_row = r
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(vertical="center", wrap_text=True,
                                   horizontal="right" if i >= 3 else "left")
        cell.border = Border(bottom=thin_blue)
    ws.row_dimensions[r].height = 26
    r += 1

    data_rows = [row for row in rows if row["start"] and row["end"] and row["end"] >= row["start"]]
    for idx, row in enumerate(data_rows):
        period = (f"{fmt_date(row['clamped'][0])} t/m {fmt_date(row['clamped'][1])}"
                  if row["clamped"] else "Buiten venster")
        vals = [fmt_date(row["start"]), fmt_date(row["end"]), row["monthly"],
                period, round(row["eligible_gross"], 2), round(row["budget"], 2)]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = Font(name="Arial", size=10, color=INK)
            cell.alignment = Alignment(vertical="center",
                                       horizontal="right" if i >= 3 else "left")
            if i == 3 or i >= 5:
                cell.number_format = euro_fmt
            if idx % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
            cell.border = Border(bottom=hair)
        r += 1
    r += 1

    totals = [
        ("Totaal meetellend brutoloon", round(total_eligible, 2), False),
        ("Totaal studiebudget (2%)", round(total_budget, 2), True),
        ("   waarvan reeds opgebouwd (t/m vandaag)", round(accrued, 2), False),
        ("   waarvan vooraf op te nemen (toekomst)", round(advance, 2), False),
    ]
    if used > 0:
        totals.append(("Al gebruikt", round(used, 2), False))
        totals.append(("Resterend", round(remaining, 2), True))

    for label, val, highlight in totals:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        lc = ws.cell(row=r, column=1, value=label)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.font = Font(name="Arial", size=10, bold=highlight,
                       color=BLUE_DEEP if highlight else INK)
        ec = ws.cell(row=r, column=5, value=val)
        ec.number_format = euro_fmt
        ec.alignment = Alignment(horizontal="right", vertical="center")
        ec.font = Font(name="Arial", size=12 if highlight else 10, bold=highlight,
                       color=BLUE_DEEP if highlight else INK)
        if highlight:
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT)
        r += 1

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    note = ws.cell(row=r, column=1,
                   value="Het bedrag 'vooraf op te nemen' betreft toekomstig "
                         "contractloon en dient bij eerder vertrek te worden terugbetaald.")
    note.font = Font(name="Arial", size=9, italic=True, color=GREY)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---- Resultaten ----------------------------------------------------------
if calculate:
    rows, contract_end, window_start = build_rows(edited)

    if window_start and contract_end:
        st.info(f"Het venster loopt van **{fmt_date(window_start)}** "
                f"t/m **{fmt_date(contract_end)}**.")

    total_eligible = sum(r["eligible_gross"] for r in rows)
    total_budget = sum(r["budget"] for r in rows)
    accrued = sum(r["past_gross"] * RATE for r in rows)
    advance = sum(r["future_gross"] * RATE for r in rows)
    used = float(already_used or 0)
    remaining = total_budget - used
    overspent = remaining < -0.005

    has_result = any(r["budget"] > 0 for r in rows)

    if not has_result:
        st.warning("Geen geldige perioden gevonden. Controleer de start-/einddatums "
                   "en het bruto maandsalaris.")
    else:
        st.subheader(f"Resultaat{' voor ' + name if name else ''}")

        c1, c2, c3 = st.columns(3)
        c1.metric("Totaal brutoloon periode", fmt_eur(total_eligible))
        c2.metric("Totaal budget (2%)", fmt_eur(total_budget))
        c3.metric("Resterend" if used > 0 else "Beschikbaar", fmt_eur(remaining),
                  delta=f"−{fmt_eur(used)}" if used > 0 else None,
                  delta_color="inverse")

        if advance > 0.005:
            d1, d2 = st.columns(2)
            d1.metric(f"Reeds opgebouwd t/m {fmt_date(today)}", fmt_eur(accrued))
            d2.metric("Vooraf op te nemen (toekomstig contract)", fmt_eur(advance),
                      help="Terug te betalen bij eerder vertrek")

        if overspent:
            st.error(f"⚠️ Deze medewerker heeft {fmt_eur(used - total_budget)} méér "
                     f"gebruikt dan het beschikbare budget.")

        # Notities per regel
        for r in rows:
            if r["note"]:
                if "Buiten" in r["note"] or "vóór de start" in r["note"]:
                    st.warning(r["note"])
                else:
                    st.info(r["note"])

        # Specificatie
        spec = []
        for r in rows:
            if r["budget"] > 0:
                spec.append({
                    "Periode": f"{fmt_date(r['clamped'][0])} – {fmt_date(r['clamped'][1])}",
                    "Maandsalaris": fmt_eur(r["monthly"]),
                    "Meetellend brutoloon": fmt_eur(r["eligible_gross"]),
                    "Budget (2%)": fmt_eur(r["budget"]),
                })
        if spec:
            st.markdown("**Specificatie**")
            st.dataframe(pd.DataFrame(spec), use_container_width=True, hide_index=True)

        # Excel-export
        excel_buf = make_excel(rows, contract_end, window_start, total_eligible,
                               total_budget, accrued, advance, used, remaining, name)
        safe = "".join(ch if ch.isalnum() else "_" for ch in (name or "medewerker"))
        st.download_button(
            "⬇️ Exporteer naar Excel",
            data=excel_buf,
            file_name=f"studiebudget_{safe}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
